import glob
import openpyxl
import os
import re
import shutil
import time

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from RPA.Browser.Selenium import Selenium
from RPA.Desktop import Desktop
from RPA.Excel.Files import Files
from RPA.FileSystem import FileSystem
from RPA.PDF import PDF
from RPA.Tables import Tables

# ===== the below parameters can be changed =====
WORKBOOK_NAME = 'workbook.xlsx'
SHEET_GENERAL = 'Agencies'
BOT_TASK = 'config.txt'
OUTPUT_DIR = os.path.join(os.getcwd(), 'output')

# ===== Do not change the following below =======
BASE_URL = "https://itdashboard.gov/"
DIVE_IN = "css:a.btn-lg-2x"
DATA_FEEDS_URL = "css:div.tuck-5 > p > a[href='/drupal/data/datafeeds']"
AGENCY_TITLES = "css:#agency-tiles-2-widget > div > div > div > div > div > div > div:nth-child(2) > a > span.h4.w200"
AGENCY_SPENDINGS = "css:#agency-tiles-2-widget > div > div > div > div > div > div > div:nth-child(2) > a > span.h1.w900"

# the below preferences might be required depending on browser saving settings
prefs = {
    "download.default_directory": OUTPUT_DIR,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True
}
browser = Selenium()
desktop = Desktop()
tables = Tables() 
# table = Tables()
excel = Files()
pdf = PDF()
files = Files()
uii_acronym_dict = {}


def get_extended_info(item):
    INVESTMENTS_TABLE = "css:#investments-table-object"
    TABLE_OBJECT_NEXT = "css:#investments-table-object_next.paginate_button.next.disabled"
    INVESTMENT_DROPDOWN = "css:select.form-control:nth-child(1)"
    INVESTMENT_SHOW_ALL = "css:select.form-control:nth-child(1) > option:last-child"
    TABLE_HEADER = "css:.dataTables_scrollHeadInner > table > thead > tr:nth-child(2)"
    TABLE_ROWS = "css:#investments-table-object.datasource-table.usa-table-borderless.dataTable.no-footer tbody tr"
    INVESTMENT_UII = "css:tr > td > a"
    # DOWNLOAD_ELEMENT = "#business-case-pdf > a"
    DOWNLOAD_BUTTON = "link:Download Business Case PDF"
    GEN_TEXT = "Generating PDF"

    global uii_acronym_dict

    print('Getting extended info on ', item)
    browser.click_element('partial link:'+item)
    browser.wait_until_element_is_visible(INVESTMENTS_TABLE, timeout=10)
    browser.click_element(INVESTMENT_DROPDOWN)
    browser.click_element(INVESTMENT_SHOW_ALL)
    browser.wait_until_element_is_visible(TABLE_OBJECT_NEXT, timeout=15)

    table_data = []
    for i in browser.get_webelements(TABLE_HEADER):
        table_data.append(re.sub(r"\['|']|\\n\\r", "", i.get_attribute('innerText'), flags=re.MULTILINE).split('\t'))
    for i in browser.get_webelements(TABLE_ROWS):
        line = re.sub(r"\['|']|\\n\\r", "", i.get_attribute('innerText'), flags=re.MULTILINE).split('\t')
        table_data.append(line) 
        uii_res = re.search(r"(\d{3})-\d+", i.get_attribute('innerText'))
        if uii_res:
            uii = uii_res.group(1)
    
    acronym = ''
    for word in item.split():
        acronym += word[0]

    files.open_workbook(os.path.join(OUTPUT_DIR, WORKBOOK_NAME))
    uii_acronym_dict.update({uii: acronym})
    update_excel(acronym, table_data, WORKBOOK_NAME)

    uii_links = []
    uii_list = browser.get_webelements(INVESTMENT_UII)
    for i in range(len(uii_list)):
        try:
            href = browser.get_webelement("css:tr:nth-child("+str(i+1)+") > td > a").get_attribute('href') 
        except:
            continue
        else:
            uii_links.append(href) 

    for url in uii_links:
        browser.go_to(url)
        try:
            browser.wait_until_element_is_visible(DOWNLOAD_BUTTON)
        except:
            print("No Download button is available at: ", url)
        else:
            browser.click_element(DOWNLOAD_BUTTON)
            browser.wait_until_element_is_not_visible(GEN_TEXT)
            time.sleep(7)
        
    time.sleep(5)  # to finish all downloads before closing the browser


def get_agencies_info():
    print("Getting the list of displayed Agencies")
    agencies_titles = browser.get_webelements(AGENCY_TITLES)
    agencies_spendings = browser.get_webelements(AGENCY_SPENDINGS)

    agencies_list = []
    for agency in agencies_titles:
        agencies_list.append(browser.get_text(agency)) 

    spendings_list = []
    for spending in agencies_spendings:
        spendings_list.append(browser.get_text(spending))

    agencies_data = list(zip(agencies_list, spendings_list))
    agencies_table = tables.create_table(data=agencies_data)
    files.create_workbook()
    update_excel(SHEET_GENERAL, agencies_table, WORKBOOK_NAME)

    return agencies_list


def update_excel(sheet_name, table_name, file_name):
    files.create_worksheet(sheet_name)
    files.set_active_worksheet(sheet_name)
    files.append_rows_to_worksheet(content=table_name)
    files.save_workbook(os.path.join(OUTPUT_DIR, file_name))


def analyse_pdf():
    collisions = 0
    uii_name_collisions = {}     # found in PDF but not in Excel
    uii_column = 1
    inv_column = 3
    global uii_acronym_dict

    highlight_green = 'BBFFAA'  # light green used for matching UII - Investment Title values
    # highlight_pink = 'FFAADD'   # light pink  used for non-matching UII - Investment Title values

    targetFile = os.path.join(OUTPUT_DIR, 'workbook.xlsx')
    wb = openpyxl.load_workbook(targetFile)
    os.chdir(OUTPUT_DIR)
    for pdf_file in glob.glob("*.pdf"):
        print("Processing ", pdf_file)
        pdf.extract_pages_from_pdf(source_path=pdf_file, output_path="page.pdf", pages=1)
        text_dict = pdf.get_text_from_pdf("page.pdf")
        text = text_dict[1]
        paragraphs = text.split('Section', 2)
        title = re.search(r"Name of this Investment: ([a-z A-Z]+.*?)2\.", paragraphs[1])
        if title:
            invest_title = title.group(1)
        else:
            print("'Name of this Investment' value was not found in ", pdf_file)
            collisions += 1
            uii_name_collisions.update({collisions: "Name of this Investment value was not found in PDF"})
        title = re.search(r"Unique Investment Identifier.*?(\d{3}-\d+)", paragraphs[1])
        if title:
            invest_uii = title.group(1)
        else:
            print("'Unique Investment Identifier' value was not found in ", pdf_file)
            collisions += 1
            uii_name_collisions.update({collisions: "UII value was not found in PDF"})

        tab = uii_acronym_dict[invest_uii[:3]]
        activeSheet = wb[tab]
        rows = activeSheet.max_row
        item_found = False
        for i in range(2, rows+1):
            #print(i, activeSheet.cell(row=i, column=uii_column).value, activeSheet.cell(row=i, column=inv_column).value)
            if invest_uii == activeSheet.cell(row=i, column=uii_column).value and invest_title == activeSheet.cell(row=i, column=inv_column).value:
                activeSheet.cell(row=i, column=uii_column).fill = PatternFill(fgColor=highlight_green, fill_type = 'solid')
                activeSheet.cell(row=i, column=inv_column).fill = PatternFill(fgColor=highlight_green, fill_type = 'solid')
                item_found = True

        if not item_found:
            uii_name_collisions.update({invest_uii: invest_title})

    for k, v in uii_name_collisions.items():
        activeSheet.append(("Collision", k, v))

    wb.save(filename=targetFile)


def main():
    if os.path.exists(OUTPUT_DIR):
        try:
            shutil.rmtree(OUTPUT_DIR)
        except OSError as e:
            print ("Error: %s - %s." % (e.filename, e.strerror))
    
    os.makedirs(OUTPUT_DIR)
    browser.set_download_directory(OUTPUT_DIR, download_pdf=True)
    browser.open_available_browser(BASE_URL, maximized=True)
    browser.wait_until_element_is_visible(AGENCY_TITLES)
    # extract summary data on agencies and spending
    agencies = get_agencies_info()
    # extract the data on particular agency and add to the Excel file
    with open(BOT_TASK, "r") as f:
        os.chdir(OUTPUT_DIR)
        for line in f:
            agency = line.strip()
            if agency:
                if not re.match(r'(^(\s+)?#)', agency):
                    if agency in agencies:
                        browser.go_to(BASE_URL)
                        browser.wait_until_element_is_visible(AGENCY_TITLES)
                        get_extended_info(agency)
                        # analyse_pdf()  
                        # agency_dir = os.path.join(OUTPUT_DIR, agency)
                        # os.makedirs(agency_dir)
                        # for pdf_file in glob.glob("*.pdf"):
                        #     if not pdf_file == 'page.pdf':
                        #         shutil.move(pdf_file, os.path.join(agency_dir, pdf_file))
                    else:
                        print("The agency '%s' was not found in the list" % agency)
    
    browser.close_browser()
    # analysing downloaded PDF files and extract the data
    analyse_pdf()
    print('Done')


if __name__ == "__main__":
    main()
    